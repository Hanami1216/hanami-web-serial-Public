import json
from openpyxl import load_workbook

wb = load_workbook('CMD_回归测试核对表.xlsx')
ws1 = wb[wb.sheetnames[0]]
ws2 = wb[wb.sheetnames[1]]

# Find the last used row in Sheet2
last_row = 1
for r in range(1, 50):
    a = ws2.cell(row=r, column=1).value
    b = ws2.cell(row=r, column=2).value
    if a or b:
        last_row = r

print(f'Last used row in Sheet2: {last_row}')
start = last_row + 2

# Load JSON to get dangerous command details
with open('ble-cmd-test-2026-07-17T01-49-27.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

by_cmd = {d['cmd']: d for d in data['details']}

danger_cmds = [
    (0x10, 'EQ_VOL_RESET'),
    (0x27, 'EQ_VOL_SAVE'),
    (0x28, 'EQ_MIC_RESET'),
    (0x37, 'EQ_MIC_SAVE'),
    (0x57, 'LIGHT_SAVE'),
    (0x6F, 'TEXT_SAVE'),
]

ws2.cell(row=start, column=1).value = '危险命令清单（需 includeDangerous: true）'
start += 1

headers = ['CMD值', '命令名称', '功能组', '跳过原因']
for i, h in enumerate(headers):
    ws2.cell(row=start, column=i+1).value = h
start += 1

for cmd_val, name in danger_cmds:
    detail = by_cmd.get(cmd_val, {})
    ws2.cell(row=start, column=1).value = f"0x{cmd_val:02X} ({cmd_val})"
    ws2.cell(row=start, column=2).value = name
    ws2.cell(row=start, column=3).value = detail.get('reason', '')
    # Check if also not supported
    ws2.cell(row=start, column=4).value = detail.get('reason', '')
    start += 1

# Also check if any of these were hardware-unsupported too
print('Dangerous commands written to Sheet2')

wb.save('CMD_回归测试核对表.xlsx')
print('Saved')
